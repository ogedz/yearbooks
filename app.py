import os
import re
import json
import sqlite3
import base64
from datetime import datetime
from flask import Flask, render_template, request, jsonify, g, abort, send_from_directory
from werkzeug.utils import secure_filename

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'yearbook.db')

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 60 * 1024 * 1024  # 60MB upload cap (multi-photo batches)

# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.executescript(open(os.path.join(BASE_DIR, 'schema.sql')).read())
    db.commit()
    db.close()


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------

PLACEHOLDER_VALUES = {'', 'nothing', 'nil', 'none', 'n/a', 'na', '-', '—', 'nill'}


def clean_text(val):
    """Normalize a free-text form answer: strip, collapse blank/placeholder to ''."""
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    s = str(val).strip()
    if s.lower() in PLACEHOLDER_VALUES:
        return ''
    s = re.sub(r'\s+', ' ', s)
    return s


def title_name(name):
    """Title-case a full name while respecting common Nigerian name casing."""
    name = clean_text(name)
    if not name:
        return ''
    parts = name.split(' ')
    fixed = []
    for p in parts:
        if p.isupper() or p.islower():
            fixed.append(p.capitalize())
        else:
            fixed.append(p)
    return ' '.join(fixed)


def slugify(text):
    text = re.sub(r'[^a-zA-Z0-9]+', '-', text.lower()).strip('-')
    return text or 'item'


def file_to_base64(file_storage):
    data = file_storage.read()
    mime = file_storage.mimetype or 'image/jpeg'
    b64 = base64.b64encode(data).decode('ascii')
    return f"data:{mime};base64,{b64}"


def row_to_dict(row):
    return dict(row) if row else None


# ---------------------------------------------------------------------------
# Public routes
# ---------------------------------------------------------------------------

@app.route('/')
def home():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    student_count = db.execute("SELECT COUNT(*) c FROM students").fetchone()['c']
    teaching_count = db.execute("SELECT COUNT(*) c FROM staff WHERE category='Teaching'").fetchone()['c']
    nonteaching_count = db.execute("SELECT COUNT(*) c FROM staff WHERE category='Non-Teaching'").fetchone()['c']
    class_count = db.execute("SELECT COUNT(DISTINCT class_name) c FROM students").fetchone()['c']
    leaders = db.execute("SELECT * FROM leadership ORDER BY sort_order ASC").fetchall()
    return render_template('home.html', settings=settings, student_count=student_count,
                            teaching_count=teaching_count, nonteaching_count=nonteaching_count,
                            class_count=class_count, leaders=leaders)


@app.route('/leadership')
def leadership():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    leaders = db.execute("SELECT * FROM leadership ORDER BY sort_order ASC").fetchall()
    return render_template('leadership.html', settings=settings, leaders=leaders)


@app.route('/staff')
def staff():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    rows = db.execute("SELECT * FROM staff ORDER BY category, department, name").fetchall()
    departments = {}
    for r in rows:
        dept = r['department'] or 'General'
        departments.setdefault(r['category'], {}).setdefault(dept, []).append(r)
    return render_template('staff.html', settings=settings, departments=departments)


@app.route('/classes')
def classes():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    rows = db.execute("""SELECT class_name, COUNT(*) c FROM students
                          GROUP BY class_name ORDER BY class_name""").fetchall()
    return render_template('classes.html', settings=settings, class_rows=rows)


@app.route('/class/<class_name>')
def class_students(class_name):
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    students = db.execute("""SELECT * FROM students WHERE class_name=? ORDER BY full_name""",
                           (class_name,)).fetchall()
    if not students:
        abort(404)
    # attach cover photo
    enriched = []
    for s in students:
        sd = dict(s)
        photo = db.execute("""SELECT data_url FROM student_photos WHERE student_id=?
                               ORDER BY is_cover DESC, sort_order ASC LIMIT 1""", (s['id'],)).fetchone()
        sd['cover_photo'] = photo['data_url'] if photo else None
        enriched.append(sd)
    return render_template('class_students.html', settings=settings, class_name=class_name, students=enriched)


@app.route('/student/<int:student_id>')
def student_profile(student_id):
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    student = db.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    if not student:
        abort(404)
    photos = db.execute("""SELECT * FROM student_photos WHERE student_id=?
                            ORDER BY is_cover DESC, sort_order ASC""", (student_id,)).fetchall()
    return render_template('student_profile.html', settings=settings, s=student, photos=photos)


@app.route('/articles')
def articles():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    rows = db.execute("""SELECT * FROM articles WHERE status='Published'
                          ORDER BY created_at DESC""").fetchall()
    return render_template('articles.html', settings=settings, articles=rows)


@app.route('/articles/<int:article_id>')
def article_detail(article_id):
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    article = db.execute("SELECT * FROM articles WHERE id=? AND status='Published'", (article_id,)).fetchone()
    if not article:
        abort(404)
    return render_template('article_detail.html', settings=settings, a=article)


@app.route('/gallery')
def gallery():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    category = request.args.get('category', 'All')
    if category and category != 'All':
        rows = db.execute("SELECT * FROM gallery_photos WHERE category=? ORDER BY created_at DESC",
                           (category,)).fetchall()
    else:
        rows = db.execute("SELECT * FROM gallery_photos ORDER BY created_at DESC").fetchall()
    cats = db.execute("SELECT DISTINCT category FROM gallery_photos ORDER BY category").fetchall()
    return render_template('gallery.html', settings=settings, photos=rows, categories=cats, active_cat=category)


@app.route('/search')
def search():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    q = request.args.get('q', '').strip()
    results = {'students': [], 'staff': [], 'articles': [], 'gallery': []}
    if q:
        like = f"%{q}%"
        results['students'] = db.execute("""SELECT * FROM students WHERE
            full_name LIKE ? OR nickname LIKE ? OR future_ambition LIKE ? OR class_name LIKE ?
            LIMIT 30""", (like, like, like, like)).fetchall()
        results['staff'] = db.execute("""SELECT * FROM staff WHERE
            name LIKE ? OR department LIKE ? OR subject LIKE ? LIMIT 30""", (like, like, like)).fetchall()
        results['articles'] = db.execute("""SELECT * FROM articles WHERE status='Published' AND
            (title LIKE ? OR author LIKE ? OR content LIKE ?) LIMIT 30""", (like, like, like)).fetchall()
        results['gallery'] = db.execute("""SELECT * FROM gallery_photos WHERE
            caption LIKE ? OR category LIKE ? LIMIT 30""", (like, like)).fetchall()
    return render_template('search.html', settings=settings, q=q, results=results)


# ---------------------------------------------------------------------------
# Admin routes (no auth — intentionally unlinked from public nav)
# ---------------------------------------------------------------------------

@app.route('/admin')
def admin_dashboard():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    student_count = db.execute("SELECT COUNT(*) c FROM students").fetchone()['c']
    staff_count = db.execute("SELECT COUNT(*) c FROM staff").fetchone()['c']
    article_count = db.execute("SELECT COUNT(*) c FROM articles").fetchone()['c']
    gallery_count = db.execute("SELECT COUNT(*) c FROM gallery_photos").fetchone()['c']
    photos_done = db.execute("""SELECT COUNT(DISTINCT student_id) c FROM student_photos""").fetchone()['c']
    classes_rows = db.execute("SELECT DISTINCT class_name FROM students ORDER BY class_name").fetchall()
    return render_template('admin/dashboard.html', settings=settings, student_count=student_count,
                            staff_count=staff_count, article_count=article_count,
                            gallery_count=gallery_count, photos_done=photos_done,
                            classes_rows=classes_rows)


@app.route('/admin/students')
def admin_students():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    class_filter = request.args.get('class', '')
    if class_filter:
        rows = db.execute("SELECT * FROM students WHERE class_name=? ORDER BY full_name", (class_filter,)).fetchall()
    else:
        rows = db.execute("SELECT * FROM students ORDER BY class_name, full_name").fetchall()
    enriched = []
    for s in rows:
        sd = dict(s)
        sd['photo_count'] = db.execute("SELECT COUNT(*) c FROM student_photos WHERE student_id=?",
                                        (s['id'],)).fetchone()['c']
        enriched.append(sd)
    classes_rows = db.execute("SELECT DISTINCT class_name FROM students ORDER BY class_name").fetchall()
    return render_template('admin/students.html', settings=settings, students=enriched,
                            classes_rows=classes_rows, class_filter=class_filter)


@app.route('/admin/students/import', methods=['POST'])
def admin_import_students():
    """Import a Google-Forms style Excel export for a given class."""
    from openpyxl import load_workbook
    import io

    file = request.files.get('file')
    class_name = clean_text(request.form.get('class_name'))
    if not file or not class_name:
        return jsonify({'ok': False, 'error': 'File and class name are required'}), 400

    wb = load_workbook(io.BytesIO(file.read()), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

    COLMAP = {
        'fullName': ['full name'],
        'nickname': ['soubriquet', 'nickname'],
        'postHeld': ['post held'],
        'dob': ['date of birth'],
        'yearOfAdmission': ['year of admission'],
        'religion': ['religion'],
        'modeOfSchooling': ['mode of schooling'],
        'stateOfOrigin': ['state of origin'],
        'homeTown': ['home town'],
        'contactNumber': ['contact number'],
        'socialMedia': ['social media'],
        'roleModel': ['role model'],
        'hobbies': ['hobbies'],
        'bestFriends': ['best friend'],
        'folksNeverForget': ['folks i will never forget'],
        'favouriteColour': ['favourite colour'],
        'favouriteSubject': ['favourite subject'],
        'toughestSubject': ['toughest subject'],
        'favouriteTeacher': ['favourite teacher'],
        'mostAdmiredTeacher': ['most admired teacher'],
        'favouriteNonTeaching': ['favourite non-teaching'],
        'mostAdmiredClassmate': ['most admired classmate'],
        'bestSeatPartner': ['best seat partner'],
        'dayOneFriend': ['day one friend'],
        'classCrush': ['class crush'],
        'teachersMissed': ['teachers that i miss'],
        'favouriteFood': ['favourite food'],
        'favouriteFruit': ['favourite fruit'],
        'movieType': ['favourite movie type'],
        'bestArtist': ['best artist'],
        'bestSlang': ['best slang'],
        'favouriteQuote': ['favourite quote'],
        'bestMoment': ['best moment'],
        'embarrassingMoment': ['most embarrassing moment'],
        'weirdMoment': ['weird moment'],
        'worstMoment': ['worst moment'],
        'neverForget': ['what you will never forget about this set'],
        'betterThanThought': ['better than i thought'],
        'impactOfUmic': ['impact(s) of umic'],
        'journeyDescription': ['describe your journey'],
        'adviceToFuture': ['advice to your future self'],
        'neverMiss': ['one thing you will never miss'],
        'secretlyMiss': ['one thing you will secretly miss'],
        'classEnjoyedMost': ['what class did you enjoy most'],
        'notablePhobia': ['notable phobia'],
        'futureAmbition': ['future ambition'],
        'dreamland': ['dreamland'],
        'peopleToMeet': ['people you would love to meet'],
        'wish': ['i wish i could'],
        'partingWords': ['parting words'],
    }

    def find_col(keywords):
        for i, h in enumerate(headers):
            hl = h.lower()
            if any(kw in hl for kw in keywords):
                return i
        return None

    field_indices = {field: find_col(kws) for field, kws in COLMAP.items()}

    raw_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        record = {}
        for field, idx in field_indices.items():
            record[field] = clean_text(row[idx]) if idx is not None and idx < len(row) else ''
        if not record.get('fullName'):
            continue
        record['fullName'] = title_name(record['fullName'])
        raw_rows.append(record)

    # Dedupe: keep the most "complete" (most non-empty fields) row per normalized name
    best_by_name = {}
    for rec in raw_rows:
        key = rec['fullName'].lower()
        score = sum(1 for v in rec.values() if v)
        if key not in best_by_name or score > best_by_name[key][0]:
            best_by_name[key] = (score, rec)

    db = get_db()
    inserted, skipped = 0, 0
    for score, rec in best_by_name.values():
        existing = db.execute("SELECT id FROM students WHERE LOWER(full_name)=? AND class_name=?",
                               (rec['fullName'].lower(), class_name)).fetchone()
        if existing:
            skipped += 1
            continue
        cols = ['full_name', 'class_name'] + list(COLMAP.keys())
        cols.remove('fullName')
        values = [rec['fullName'], class_name] + [rec.get(k, '') for k in COLMAP.keys() if k != 'fullName']
        placeholders = ','.join(['?'] * len(values))
        colnames = ','.join(['full_name', 'class_name'] + [to_snake(k) for k in COLMAP.keys() if k != 'fullName'])
        db.execute(f"INSERT INTO students ({colnames}) VALUES ({placeholders})", values)
        inserted += 1
    db.commit()
    return jsonify({'ok': True, 'inserted': inserted, 'skipped_duplicates': skipped, 'total_seen': len(raw_rows)})


def to_snake(camel):
    return re.sub(r'(?<!^)(?=[A-Z])', '_', camel).lower()


@app.route('/admin/students/<int:student_id>/edit', methods=['POST'])
def admin_edit_student(student_id):
    db = get_db()
    data = request.get_json(force=True)
    allowed = {to_snake(k) for k in [
        'fullName','nickname','postHeld','dob','yearOfAdmission','religion','modeOfSchooling',
        'stateOfOrigin','homeTown','contactNumber','socialMedia','roleModel','hobbies','bestFriends',
        'folksNeverForget','favouriteColour','favouriteSubject','toughestSubject','favouriteTeacher',
        'mostAdmiredTeacher','favouriteNonTeaching','mostAdmiredClassmate','bestSeatPartner',
        'dayOneFriend','classCrush','teachersMissed','favouriteFood','favouriteFruit','movieType',
        'bestArtist','bestSlang','favouriteQuote','bestMoment','embarrassingMoment','weirdMoment',
        'worstMoment','neverForget','betterThanThought','impactOfUmic','journeyDescription',
        'adviceToFuture','neverMiss','secretlyMiss','classEnjoyedMost','notablePhobia','futureAmbition',
        'dreamland','peopleToMeet','wish','partingWords','className'
    ]}
    sets, vals = [], []
    for k, v in data.items():
        snake = to_snake(k)
        if snake in allowed or snake == 'class_name':
            sets.append(f"{snake}=?")
            vals.append(clean_text(v) if snake != 'class_name' else v)
    if not sets:
        return jsonify({'ok': False, 'error': 'No valid fields'}), 400
    vals.append(student_id)
    db.execute(f"UPDATE students SET {', '.join(sets)} WHERE id=?", vals)
    db.commit()
    return jsonify({'ok': True})


@app.route('/admin/students/<int:student_id>/delete', methods=['POST'])
def admin_delete_student(student_id):
    db = get_db()
    db.execute("DELETE FROM student_photos WHERE student_id=?", (student_id,))
    db.execute("DELETE FROM students WHERE id=?", (student_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/admin/students/<int:student_id>/photos-list')
def admin_student_photos_list(student_id):
    db = get_db()
    rows = db.execute("""SELECT id, data_url, is_cover FROM student_photos WHERE student_id=?
                          ORDER BY is_cover DESC, sort_order ASC""", (student_id,)).fetchall()
    return jsonify({'photos': [dict(r) for r in rows]})


@app.route('/admin/students/<int:student_id>/photos', methods=['POST'])
def admin_upload_student_photos(student_id):
    db = get_db()
    student = db.execute("SELECT id FROM students WHERE id=?", (student_id,)).fetchone()
    if not student:
        return jsonify({'ok': False, 'error': 'Student not found'}), 404
    files = request.files.getlist('photos')
    if not files:
        return jsonify({'ok': False, 'error': 'No photos provided'}), 400
    existing_count = db.execute("SELECT COUNT(*) c FROM student_photos WHERE student_id=?",
                                 (student_id,)).fetchone()['c']
    has_cover = db.execute("SELECT COUNT(*) c FROM student_photos WHERE student_id=? AND is_cover=1",
                            (student_id,)).fetchone()['c'] > 0
    added = 0
    for i, f in enumerate(files):
        if not f or not f.filename:
            continue
        data_url = file_to_base64(f)
        is_cover = 1 if (not has_cover and i == 0 and existing_count == 0) else 0
        db.execute("""INSERT INTO student_photos (student_id, data_url, is_cover, sort_order)
                      VALUES (?, ?, ?, ?)""", (student_id, data_url, is_cover, existing_count + i))
        added += 1
    db.commit()
    return jsonify({'ok': True, 'added': added})


@app.route('/admin/photos/<int:photo_id>/set-cover', methods=['POST'])
def admin_set_cover_photo(photo_id):
    db = get_db()
    photo = db.execute("SELECT student_id FROM student_photos WHERE id=?", (photo_id,)).fetchone()
    if not photo:
        return jsonify({'ok': False, 'error': 'Photo not found'}), 404
    db.execute("UPDATE student_photos SET is_cover=0 WHERE student_id=?", (photo['student_id'],))
    db.execute("UPDATE student_photos SET is_cover=1 WHERE id=?", (photo_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/admin/photos/<int:photo_id>/delete', methods=['POST'])
def admin_delete_student_photo(photo_id):
    db = get_db()
    db.execute("DELETE FROM student_photos WHERE id=?", (photo_id,))
    db.commit()
    return jsonify({'ok': True})


# --- Staff admin ---

@app.route('/admin/staff')
def admin_staff():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    rows = db.execute("SELECT * FROM staff ORDER BY category, department, name").fetchall()
    return render_template('admin/staff.html', settings=settings, staff=rows)


@app.route('/admin/staff/add', methods=['POST'])
def admin_add_staff():
    db = get_db()
    f = request.form
    photo_url = ''
    if 'photo' in request.files and request.files['photo'].filename:
        photo_url = file_to_base64(request.files['photo'])
    db.execute("""INSERT INTO staff (name, department, subject, phone, category, years_of_service, quote, photo)
                  VALUES (?,?,?,?,?,?,?,?)""",
               (clean_text(f.get('name')), clean_text(f.get('department')), clean_text(f.get('subject')),
                clean_text(f.get('phone')), f.get('category', 'Teaching'), clean_text(f.get('years_of_service')),
                clean_text(f.get('quote')), photo_url))
    db.commit()
    return jsonify({'ok': True})


@app.route('/admin/staff/<int:staff_id>/delete', methods=['POST'])
def admin_delete_staff(staff_id):
    db = get_db()
    db.execute("DELETE FROM staff WHERE id=?", (staff_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/admin/staff/import', methods=['POST'])
def admin_import_staff():
    from openpyxl import load_workbook
    import io
    file = request.files.get('file')
    if not file:
        return jsonify({'ok': False, 'error': 'File required'}), 400
    wb = load_workbook(io.BytesIO(file.read()), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]

    def find_col(keywords):
        for i, h in enumerate(headers):
            if any(kw in h for kw in keywords):
                return i
        return None

    idx_name = find_col(['name'])
    idx_dept = find_col(['department'])
    idx_subject = find_col(['subject'])
    idx_phone = find_col(['phone'])
    idx_type = find_col(['type', 'category'])
    idx_years = find_col(['years'])
    idx_quote = find_col(['quote'])

    db = get_db()
    added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        name = clean_text(row[idx_name]) if idx_name is not None else ''
        if not name:
            continue
        name = title_name(name)
        category = clean_text(row[idx_type]) if idx_type is not None else 'Teaching'
        category = 'Non-Teaching' if 'non' in category.lower() else 'Teaching'
        existing = db.execute("SELECT id FROM staff WHERE LOWER(name)=?", (name.lower(),)).fetchone()
        if existing:
            continue
        db.execute("""INSERT INTO staff (name, department, subject, phone, category, years_of_service, quote, photo)
                      VALUES (?,?,?,?,?,?,?,'')""",
                   (name,
                    clean_text(row[idx_dept]) if idx_dept is not None else '',
                    clean_text(row[idx_subject]) if idx_subject is not None else '',
                    clean_text(row[idx_phone]) if idx_phone is not None else '',
                    category,
                    clean_text(row[idx_years]) if idx_years is not None else '',
                    clean_text(row[idx_quote]) if idx_quote is not None else ''))
        added += 1
    db.commit()
    return jsonify({'ok': True, 'added': added})


@app.route('/admin/staff/<int:staff_id>/photo', methods=['POST'])
def admin_upload_staff_photo(staff_id):
    db = get_db()
    f = request.files.get('photo')
    if not f:
        return jsonify({'ok': False, 'error': 'No photo'}), 400
    data_url = file_to_base64(f)
    db.execute("UPDATE staff SET photo=? WHERE id=?", (data_url, staff_id))
    db.commit()
    return jsonify({'ok': True})


# --- Leadership admin ---

@app.route('/admin/leadership')
def admin_leadership():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    rows = db.execute("SELECT * FROM leadership ORDER BY sort_order").fetchall()
    return render_template('admin/leadership.html', settings=settings, leaders=rows)


@app.route('/admin/leadership/add', methods=['POST'])
def admin_add_leader():
    db = get_db()
    f = request.form
    photo_url = ''
    if 'photo' in request.files and request.files['photo'].filename:
        photo_url = file_to_base64(request.files['photo'])
    max_order = db.execute("SELECT COALESCE(MAX(sort_order),0) m FROM leadership").fetchone()['m']
    db.execute("""INSERT INTO leadership (name, title, message, photo, sort_order)
                  VALUES (?,?,?,?,?)""",
               (clean_text(f.get('name')), clean_text(f.get('title')), clean_text(f.get('message')),
                photo_url, max_order + 1))
    db.commit()
    return jsonify({'ok': True})


@app.route('/admin/leadership/<int:leader_id>/delete', methods=['POST'])
def admin_delete_leader(leader_id):
    db = get_db()
    db.execute("DELETE FROM leadership WHERE id=?", (leader_id,))
    db.commit()
    return jsonify({'ok': True})


# --- Articles admin ---

@app.route('/admin/articles')
def admin_articles():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    rows = db.execute("SELECT * FROM articles ORDER BY created_at DESC").fetchall()
    return render_template('admin/articles.html', settings=settings, articles=rows)


@app.route('/admin/articles/add', methods=['POST'])
def admin_add_article():
    db = get_db()
    f = request.form
    image_url = ''
    if 'cover_image' in request.files and request.files['cover_image'].filename:
        image_url = file_to_base64(request.files['cover_image'])
    db.execute("""INSERT INTO articles (title, author, category, content, cover_image, status, created_at)
                  VALUES (?,?,?,?,?,?,?)""",
               (clean_text(f.get('title')), clean_text(f.get('author')), f.get('category', 'Reflection'),
                f.get('content', ''), image_url, f.get('status', 'Published'),
                datetime.utcnow().isoformat()))
    db.commit()
    return jsonify({'ok': True})


@app.route('/admin/articles/<int:article_id>/delete', methods=['POST'])
def admin_delete_article(article_id):
    db = get_db()
    db.execute("DELETE FROM articles WHERE id=?", (article_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/admin/articles/<int:article_id>/toggle', methods=['POST'])
def admin_toggle_article(article_id):
    db = get_db()
    a = db.execute("SELECT status FROM articles WHERE id=?", (article_id,)).fetchone()
    if not a:
        return jsonify({'ok': False}), 404
    new_status = 'Draft' if a['status'] == 'Published' else 'Published'
    db.execute("UPDATE articles SET status=? WHERE id=?", (new_status, article_id))
    db.commit()
    return jsonify({'ok': True, 'status': new_status})


# --- Gallery admin ---

@app.route('/admin/gallery')
def admin_gallery():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    rows = db.execute("SELECT * FROM gallery_photos ORDER BY created_at DESC").fetchall()
    return render_template('admin/gallery.html', settings=settings, photos=rows)


@app.route('/admin/gallery/upload', methods=['POST'])
def admin_upload_gallery():
    db = get_db()
    caption = clean_text(request.form.get('caption'))
    category = request.form.get('category', 'Events')
    files = request.files.getlist('photos')
    added = 0
    for f in files:
        if not f or not f.filename:
            continue
        data_url = file_to_base64(f)
        db.execute("""INSERT INTO gallery_photos (caption, category, data_url, created_at)
                      VALUES (?,?,?,?)""", (caption, category, data_url, datetime.utcnow().isoformat()))
        added += 1
    db.commit()
    return jsonify({'ok': True, 'added': added})


@app.route('/admin/gallery/<int:photo_id>/delete', methods=['POST'])
def admin_delete_gallery_photo(photo_id):
    db = get_db()
    db.execute("DELETE FROM gallery_photos WHERE id=?", (photo_id,))
    db.commit()
    return jsonify({'ok': True})


# --- Settings admin ---

@app.route('/admin/settings')
def admin_settings():
    db = get_db()
    settings = row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone())
    return render_template('admin/settings.html', settings=settings)


@app.route('/admin/settings/update', methods=['POST'])
def admin_update_settings():
    db = get_db()
    f = request.form
    logo_url = None
    if 'logo' in request.files and request.files['logo'].filename:
        logo_url = file_to_base64(request.files['logo'])
    sets, vals = [], []
    for field in ['school_name', 'motto', 'graduation_year', 'theme_primary', 'theme_accent', 'principal_message']:
        if field in f:
            sets.append(f"{field}=?")
            vals.append(clean_text(f.get(field)) if field != 'principal_message' else f.get(field, ''))
    if logo_url:
        sets.append("logo=?")
        vals.append(logo_url)
    if sets:
        db.execute(f"UPDATE settings SET {', '.join(sets)} WHERE id=1", vals)
        db.commit()
    return jsonify({'ok': True})


if not os.path.exists(DB_PATH):
    init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
